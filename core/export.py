"""
Export utilities for time series data and analysis results
"""
import json
import numpy as np
from pathlib import Path
from typing import Optional, Dict, Any
from datetime import datetime


def export_timeseries_csv(timeseries, filepath: str, include_metadata: bool = True) -> bool:
    """
    Zaman serisi verisini CSV formatında dışa aktar.
    
    Parameters
    ----------
    timeseries : TimeSeries
        Dışa aktarılacak zaman serisi
    filepath : str
        Hedef dosya yolu
    include_metadata : bool
        Metadata satırlarını ekle (dt, t0, etc.)
    
    Returns
    -------
    bool
        Başarılı ise True
    """
    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            # Metadata (yorum satırı olarak)
            if include_metadata:
                f.write(f"# Time Series Data Export\n")
                f.write(f"# Exported: {datetime.now().isoformat()}\n")
                f.write(f"# dt: {timeseries.dt}\n")
                f.write(f"# t0: {timeseries.t0}\n")
                f.write(f"# N: {len(timeseries.data)}\n")
                if hasattr(timeseries, 'metadata') and timeseries.metadata:
                    for key, value in timeseries.metadata.items():
                        f.write(f"# {key}: {value}\n")
                f.write("#\n")
            
            # Header
            f.write("time,value\n")
            
            # Data
            for t, val in zip(timeseries.time, timeseries.data):
                f.write(f"{t:.8f},{val:.8f}\n")
        
        return True
    except Exception as e:
        print(f"CSV export error: {e}")
        return False


def export_analysis_results_json(results: Dict[str, Any], filepath: str) -> bool:
    """
    Analiz sonuçlarını JSON formatında dışa aktar.
    
    Parameters
    ----------
    results : dict
        Analiz sonuçları dictionary'si
    filepath : str
        Hedef dosya yolu
    
    Returns
    -------
    bool
        Başarılı ise True
    """
    try:
        # NumPy array'leri list'e çevir
        serializable = _make_json_serializable(results)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(serializable, f, indent=2, ensure_ascii=False)
        
        return True
    except Exception as e:
        print(f"JSON export error: {e}")
        return False


def export_plot_png(plot_widget, filepath: str, width: int = 1920, height: int = 1080) -> bool:
    """
    PyQtGraph plot widget'ını PNG olarak dışa aktar.
    
    Parameters
    ----------
    plot_widget : PlotWidget
        PyQtGraph PlotWidget
    filepath : str
        Hedef dosya yolu
    width : int
        Görüntü genişliği (piksel)
    height : int
        Görüntü yüksekliği (piksel)
    
    Returns
    -------
    bool
        Başarılı ise True
    """
    try:
        from PySide6.QtCore import QSize
        from PySide6.QtGui import QImage, QPainter
        
        # Render widget to image
        size = QSize(width, height)
        image = QImage(size, QImage.Format_ARGB32)
        image.fill(0)  # Transparent background
        
        painter = QPainter(image)
        plot_widget.render(painter)
        painter.end()
        
        # Save
        success = image.save(filepath, "PNG")
        return success
    except Exception as e:
        print(f"PNG export error: {e}")
        return False


def _make_json_serializable(obj):
    """
    Recursive function to convert NumPy types to Python native types.
    """
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        return float(obj)
    elif isinstance(obj, dict):
        return {key: _make_json_serializable(value) for key, value in obj.items()}
    elif isinstance(obj, (list, tuple)):
        return [_make_json_serializable(item) for item in obj]
    else:
        return obj


# ---------------------------------------------------------------------------
# Generic table-style data export (data table panelinin sag-tik export'u icin)
# ---------------------------------------------------------------------------

def export_table_data_csv(headers, rows, filepath: str,
                          metadata: Optional[Dict[str, Any]] = None) -> bool:
    """
    Tablo verisini (sutun basliklari + satirlar) CSV formatinda yazar.

    metadata varsa basa "# key: value" yorum satirlari eklenir.
    """
    import csv
    try:
        with open(filepath, 'w', encoding='utf-8', newline='') as f:
            if metadata:
                for k, v in metadata.items():
                    f.write(f"# {k}: {v}\n")
                f.write("#\n")
            writer = csv.writer(f)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)
        return True
    except Exception as e:
        print(f"CSV table export error: {e}")
        return False


def export_table_data_xlsx(headers, rows, filepath: str,
                           metadata: Optional[Dict[str, Any]] = None) -> bool:
    """
    Tablo verisini XLSX (Excel) formatinda yazar.

    metadata varsa ilk satirlara "# key: value" olarak yazilir.
    Sayisal degerler float olarak saklanir (Excel formul/sort uyumu).
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        print("XLSX export requires openpyxl: pip install openpyxl")
        return False

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        row_idx = 1
        if metadata:
            for k, v in metadata.items():
                ws.cell(row=row_idx, column=1, value=f"# {k}: {v}")
                row_idx += 1
            row_idx += 1  # bos satir

        # Headers (bold)
        from openpyxl.styles import Font
        bold = Font(bold=True)
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(h))
            cell.font = bold
        row_idx += 1

        # Data rows
        for row in rows:
            for col_idx, val in enumerate(row, start=1):
                # Sayisal degerleri float olarak yaz
                if isinstance(val, (int, np.integer)):
                    ws.cell(row=row_idx, column=col_idx, value=int(val))
                elif isinstance(val, (float, np.floating)):
                    ws.cell(row=row_idx, column=col_idx, value=float(val))
                else:
                    ws.cell(row=row_idx, column=col_idx, value=str(val))
            row_idx += 1

        # Sutun genislikleri otomatik (basit heuristik)
        for col_idx, h in enumerate(headers, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, len(str(h)) + 2)

        wb.save(filepath)
        return True
    except Exception as e:
        print(f"XLSX table export error: {e}")
        return False


def export_table_data_txt(headers, rows, filepath: str,
                          metadata: Optional[Dict[str, Any]] = None,
                          separator: str = '\t') -> bool:
    """
    Tablo verisini metin dosyasina yazar (varsayilan ayrim: tab).

    metadata varsa basa "# key: value" yorum satirlari eklenir.
    """
    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            if metadata:
                for k, v in metadata.items():
                    f.write(f"# {k}: {v}\n")
                f.write("#\n")
            f.write(separator.join(str(h) for h in headers) + "\n")
            for row in rows:
                f.write(separator.join(_format_cell(v) for v in row) + "\n")
        return True
    except Exception as e:
        print(f"TXT table export error: {e}")
        return False


def _format_cell(val) -> str:
    """Hucre degerini metin olarak duzenle (float'lar 6g)."""
    if isinstance(val, (int, np.integer)):
        return str(int(val))
    if isinstance(val, (float, np.floating)):
        if np.isnan(val):
            return "nan"
        return f"{float(val):.6g}"
    return str(val)


def create_analysis_summary(timeseries, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
    """
    Tam analiz özeti oluştur (session kaydetmek için).
    
    Parameters
    ----------
    timeseries : TimeSeries
        Zaman serisi verisi
    analysis_results : dict
        Tüm analiz sonuçları
    
    Returns
    -------
    dict
        Tam analiz özeti
    """
    summary = {
        "export_timestamp": datetime.now().isoformat(),
        "timeseries": {
            "dt": timeseries.dt,
            "t0": timeseries.t0,
            "n_points": len(timeseries.data),
            "data": timeseries.data.tolist(),
            "time": timeseries.time.tolist(),
            "metadata": getattr(timeseries, 'metadata', {})
        },
        "analysis_results": _make_json_serializable(analysis_results)
    }
    
    return summary
